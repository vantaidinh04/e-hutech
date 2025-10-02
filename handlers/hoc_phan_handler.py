#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Handler xử lý học phần từ hệ thống HUTECH
"""

import json
import logging
import aiohttp
import io
from typing import Dict, Any, Optional, List
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from config.config import Config

logger = logging.getLogger(__name__)

class HocPhanHandler:
    def __init__(self, db_manager, cache_manager):
        self.db_manager = db_manager
        self.cache_manager = cache_manager
        self.config = Config()
    
    async def handle_hoc_phan(self, telegram_user_id: int) -> Dict[str, Any]:
        """
        Xử lý lấy danh sách năm học - học kỳ của người dùng
        
        Args:
            telegram_user_id: ID của người dùng trên Telegram
            
        Returns:
            Dict chứa kết quả và dữ liệu năm học - học kỳ
        """
        try:
            cache_key = f"nam_hoc_hoc_ky:{telegram_user_id}"

            # 1. Kiểm tra cache
            cached_result = await self.cache_manager.get(cache_key)
            if cached_result:
                nam_hoc_data = cached_result.get("data")
                timestamp = cached_result.get("timestamp")

                processed_data = self._process_nam_hoc_hoc_ky_data(nam_hoc_data)
                processed_data["timestamp"] = timestamp

                return {
                    "success": True,
                    "message": "Lấy danh sách năm học - học kỳ thành công",
                    "data": processed_data
                }

            # 2. Nếu cache miss, gọi API
            token = await self._get_user_token(telegram_user_id)
            
            if not token:
                return {
                    "success": False,
                    "message": "Bạn chưa đăng nhập. Vui lòng sử dụng /login để đăng nhập.",
                    "data": None
                }
            
            response_data = await self._call_nam_hoc_hoc_ky_api(token)
            
            # 3. Lưu vào cache
            if response_data and isinstance(response_data, list):
                await self.cache_manager.set(cache_key, response_data, ttl=86400) # Cache trong 24 giờ
            
            # Kiểm tra kết quả
            if response_data and isinstance(response_data, list):
                # Xử lý dữ liệu năm học - học kỳ
                processed_data = self._process_nam_hoc_hoc_ky_data(response_data)
                processed_data["timestamp"] = datetime.utcnow().isoformat() # Thêm timestamp mới
                
                return {
                    "success": True,
                    "message": "Lấy danh sách năm học - học kỳ thành công (dữ liệu mới)",
                    "data": processed_data
                }
            else:
                return {
                    "success": False,
                    "message": "🚫 *Lỗi*\n\nKhông thể lấy danh sách năm học - học kỳ. Vui lòng thử lại sau.",
                    "data": response_data
                }
        
        except Exception as e:
            logger.error(f"Học phần error for user {telegram_user_id}: {e}")
            return {
                "success": False,
                "message": f"🚫 *Lỗi*\n\nĐã xảy ra lỗi khi lấy danh sách năm học - học kỳ: {str(e)}",
                "data": None
            }
    
    async def handle_search_hoc_phan(self, telegram_user_id: int, nam_hoc_hoc_ky_list: List[str]) -> Dict[str, Any]:
        """
        Xử lý tìm kiếm học phần theo danh sách năm học - học kỳ
        
        Args:
            telegram_user_id: ID của người dùng trên Telegram
            nam_hoc_hoc_ky_list: Danh sách mã năm học - học kỳ
            
        Returns:
            Dict chứa kết quả và dữ liệu học phần
        """
        try:
            # Tạo cache key dựa trên user_id và danh sách năm học
            cache_key = f"search_hoc_phan:{telegram_user_id}:{':'.join(sorted(nam_hoc_hoc_ky_list))}"

            # 1. Kiểm tra cache
            cached_result = await self.cache_manager.get(cache_key)
            if cached_result:
                processed_data = self._process_search_hoc_phan_data(cached_result.get("data", []))
                return {
                    "success": True,
                    "message": "Tìm kiếm học phần thành công (từ cache)",
                    "data": processed_data
                }

            # 2. Nếu cache miss, gọi API
            token = await self._get_user_token(telegram_user_id)
            
            if not token:
                return {
                    "success": False,
                    "message": "Bạn chưa đăng nhập. Vui lòng sử dụng /login để đăng nhập.",
                    "data": None
                }
            
            response_data = await self._call_search_hoc_phan_api(token, nam_hoc_hoc_ky_list)
            
            # 3. Lưu vào cache nếu thành công
            if response_data and isinstance(response_data, list):
                await self.cache_manager.set(cache_key, response_data, ttl=3600) # Cache trong 1 giờ
            
            # Kiểm tra kết quả
            if response_data and isinstance(response_data, list):
                # Xử lý dữ liệu học phần
                # Xử lý dữ liệu học phần
                processed_data = self._process_search_hoc_phan_data(response_data)
                processed_data["timestamp"] = datetime.utcnow().isoformat()
                
                return {
                    "success": True,
                    "message": "Tìm kiếm học phần thành công",
                    "data": processed_data
                }
            else:
                return {
                    "success": False,
                    "message": "🚫 *Lỗi*: Không thể tìm kiếm học phần. Vui lòng thử lại sau.",
                    "data": response_data
                }
        
        except Exception as e:
            logger.error(f"Search học phần error for user {telegram_user_id}: {e}")
            return {
                "success": False,
                "message": f"🚫 *Lỗi*\n\nĐã xảy ra lỗi khi tìm kiếm học phần: {str(e)}",
                "data": None
            }
    
    async def handle_diem_danh(self, telegram_user_id: int, key_lop_hoc_phan: str) -> Dict[str, Any]:
        """
        Xử lý lấy lịch sử điểm danh của một học phần
        
        Args:
            telegram_user_id: ID của người dùng trên Telegram
            key_lop_hoc_phan: Khóa lớp học phần
            
        Returns:
            Dict chứa kết quả và dữ liệu điểm danh
        """
        try:
            cache_key = f"diem_danh:{telegram_user_id}:{key_lop_hoc_phan}"

            # 1. Kiểm tra cache
            cached_result = await self.cache_manager.get(cache_key)
            if cached_result:
                processed_data = self._process_diem_danh_data(cached_result.get("data", []))
                return {
                    "success": True,
                    "message": "Lấy lịch sử điểm danh thành công (từ cache)",
                    "data": processed_data
                }

            # 2. Nếu cache miss, gọi API
            token = await self._get_user_token(telegram_user_id)
            
            if not token:
                return {
                    "success": False,
                    "message": "Bạn chưa đăng nhập. Vui lòng sử dụng /login để đăng nhập.",
                    "data": None
                }
            
            response_data = await self._call_diem_danh_api(token, key_lop_hoc_phan)
            
            # 3. Lưu vào cache nếu thành công
            if response_data and isinstance(response_data, dict) and "result" in response_data:
                await self.cache_manager.set(cache_key, response_data["result"], ttl=3600) # Cache trong 1 giờ

            # Kiểm tra kết quả
            if response_data and isinstance(response_data, dict) and "result" in response_data:
                # Xử lý dữ liệu điểm danh
                diem_danh_list = response_data["result"]
                processed_data = self._process_diem_danh_data(diem_danh_list)
                processed_data["timestamp"] = datetime.utcnow().isoformat()
                
                return {
                    "success": True,
                    "message": "Lấy lịch sử điểm danh thành công",
                    "data": processed_data
                }
            else:
                # Xử lý lỗi từ API
                error_message = "Danh sách điểm danh chưa được cập nhật"
                if response_data and response_data.get("error"):
                    try:
                        # Thử parse message nếu nó là JSON string
                        api_error_details = json.loads(response_data.get("message", "{}"))
                        # Ưu tiên lấy message từ reasons, sau đó là errorMessage
                        extracted_message = api_error_details.get("reasons", {}).get("message") or api_error_details.get("errorMessage")
                        if extracted_message:
                             error_message = extracted_message.split(" - ", 1)[-1] # Lấy phần thông báo lỗi chính
                    except (json.JSONDecodeError, AttributeError):
                        # Nếu message không phải JSON hoặc không có cấu trúc mong đợi, sử dụng message gốc
                        if isinstance(response_data.get("message"), str):
                            error_message = response_data["message"]

                logger.warning(f"Invalid response data or API error: {response_data}")
                return {
                    "success": False,
                    "message": f"🚫 *Lỗi*\n\n{error_message}",
                    "data": response_data
                }
        
        except Exception as e:
            logger.error(f"Điểm danh error for user {telegram_user_id}: {e}", exc_info=True)
            return {
                "success": False,
                "message": f"🚫 *Lỗi*\n\nĐã xảy ra lỗi khi lấy lịch sử điểm danh: {str(e)}",
                "data": None
            }
    
    async def handle_danh_sach_sinh_vien(self, telegram_user_id: int, key_lop_hoc_phan: str) -> Dict[str, Any]:
        """
        Xử lý lấy danh sách sinh viên của một học phần
        
        Args:
            telegram_user_id: ID của người dùng trên Telegram
            key_lop_hoc_phan: Khóa lớp học phần
            
        Returns:
            Dict chứa kết quả và dữ liệu danh sách sinh viên
        """
        try:
            # Lấy token của người dùng
            token = await self._get_user_token(telegram_user_id)
            
            if not token:
                return {
                    "success": False,
                    "message": "Bạn chưa đăng nhập. Vui lòng sử dụng /login để đăng nhập.",
                    "data": None
                }
            
            # Gọi API danh sách sinh viên
            response_data = await self._call_danh_sach_sinh_vien_api(token, key_lop_hoc_phan)
            
            # Kiểm tra kết quả
            if response_data and isinstance(response_data, dict):
                # Xử lý dữ liệu danh sách sinh viên
                processed_data = self._process_danh_sach_sinh_vien_data(response_data)
                
                return {
                    "success": True,
                    "message": "Lấy danh sách sinh viên thành công",
                    "data": processed_data
                }
            else:
                return {
                    "success": False,
                    "message": "🚫 *Lỗi*\n\nKhông thể lấy danh sách sinh viên. Vui lòng thử lại sau.",
                    "data": response_data
                }
        
        except Exception as e:
            logger.error(f"Danh sách sinh viên error for user {telegram_user_id}: {e}")
            return {
                "success": False,
                "message": f"🚫 *Lỗi*\n\nĐã xảy ra lỗi khi lấy danh sách sinh viên: {str(e)}",
                "data": None
            }
    
    async def _call_nam_hoc_hoc_ky_api(self, token: str) -> Optional[Dict[str, Any]]:
        """
        Gọi API lấy danh sách năm học - học kỳ của HUTECH
        
        Args:
            token: Token xác thực
            
        Returns:
            Response data từ API hoặc None nếu có lỗi
        """
        try:
            url = f"{self.config.HUTECH_API_BASE_URL}{self.config.HUTECH_HOC_PHAN_NAM_HOC_HOC_KY_ENDPOINT}"
            
            # Tạo headers riêng cho API học phần
            headers = self.config.HUTECH_MOBILE_HEADERS.copy()
            headers["authorization"] = f"JWT {token}"
            
            async with aiohttp.ClientSession() as session:
                async with session.get(
                    url,
                    headers=headers
                ) as response:
                    if response.status == 200:
                        return await response.json()
                    else:
                        error_text = await response.text()
                        logger.error(f"Năm học - học kỳ API error: {response.status} - {error_text}")
                        return {
                            "error": True,
                            "status_code": response.status,
                            "message": error_text
                        }
        
        except aiohttp.ClientError as e:
            logger.error(f"HTTP client error: {e}")
            return {
                "error": True,
                "message": f"Lỗi kết nối: {str(e)}"
            }
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error: {e}")
            return {
                "error": True,
                "message": f"Lỗi phân tích dữ liệu: {str(e)}"
            }
        except Exception as e:
            logger.error(f"Unexpected error: {e}")
            return {
                "error": True,
                "message": f"Lỗi không xác định: {str(e)}"
            }
    
    async def _call_search_hoc_phan_api(self, token: str, nam_hoc_hoc_ky_list: List[str]) -> Optional[Dict[str, Any]]:
        """
        Gọi API tìm kiếm học phần của HUTECH
        
        Args:
            token: Token xác thực
            nam_hoc_hoc_ky_list: Danh sách mã năm học - học kỳ
            
        Returns:
            Response data từ API hoặc None nếu có lỗi
        """
        try:
            url = f"{self.config.HUTECH_API_BASE_URL}{self.config.HUTECH_HOC_PHAN_SEARCH_ENDPOINT}"
            
            # Tạo headers riêng cho API học phần
            headers = self.config.HUTECH_MOBILE_HEADERS.copy()
            headers["authorization"] = f"JWT {token}"
            
            # Tạo request body
            request_body = {
                "nam_hoc_hoc_ky": nam_hoc_hoc_ky_list
            }
            
            async with aiohttp.ClientSession() as session:
                async with session.post(
                    url,
                    headers=headers,
                    json=request_body
                ) as response:
                    if response.status == 200:
                        return await response.json()
                    else:
                        error_text = await response.text()
                        logger.error(f"Search học phần API error: {response.status} - {error_text}")
                        return {
                            "error": True,
                            "status_code": response.status,
                            "message": error_text
                        }
        
        except aiohttp.ClientError as e:
            logger.error(f"HTTP client error: {e}")
            return {
                "error": True,
                "message": f"Lỗi kết nối: {str(e)}"
            }
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error: {e}")
            return {
                "error": True,
                "message": f"Lỗi phân tích dữ liệu: {str(e)}"
            }
        except Exception as e:
            logger.error(f"Unexpected error: {e}")
            return {
                "error": True,
                "message": f"Lỗi không xác định: {str(e)}"
            }
    
    async def _call_diem_danh_api(self, token: str, key_lop_hoc_phan: str) -> Optional[Dict[str, Any]]:
        """
        Gọi API điểm danh của HUTECH
        
        Args:
            token: Token xác thực
            key_lop_hoc_phan: Khóa lớp học phần
            
        Returns:
            Response data từ API hoặc None nếu có lỗi
        """
        try:
            url = f"{self.config.HUTECH_API_BASE_URL}{self.config.HUTECH_HOC_PHAN_DIEM_DANH_ENDPOINT}"
            
            # Tạo headers riêng cho API học phần
            headers = self.config.HUTECH_MOBILE_HEADERS.copy()
            headers["authorization"] = f"JWT {token}"
            
            # Tạo query parameters
            params = {
                "key_lop_hoc_phan": key_lop_hoc_phan
            }
            
            async with aiohttp.ClientSession() as session:
                async with session.get(
                    url,
                    headers=headers,
                    params=params
                ) as response:
                    
                    if response.status == 200:
                        response_data = await response.json()
                        return response_data
                    else:
                        error_text = await response.text()
                        logger.error(f"Điểm danh API error: {response.status} - {error_text}")
                        return {
                            "error": True,
                            "status_code": response.status,
                            "message": error_text
                        }
        
        except aiohttp.ClientError as e:
            logger.error(f"HTTP client error: {e}")
            return {
                "error": True,
                "message": f"Lỗi kết nối: {str(e)}"
            }
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error: {e}")
            return {
                "error": True,
                "message": f"Lỗi phân tích dữ liệu: {str(e)}"
            }
        except Exception as e:
            logger.error(f"Unexpected error: {e}", exc_info=True)
            return {
                "error": True,
                "message": f"Lỗi không xác định: {str(e)}"
            }
    
    async def _call_danh_sach_sinh_vien_api(self, token: str, key_lop_hoc_phan: str) -> Optional[Dict[str, Any]]:
        """
        Gọi API danh sách sinh viên của HUTECH
        
        Args:
            token: Token xác thực
            key_lop_hoc_phan: Khóa lớp học phần
            
        Returns:
            Response data từ API hoặc None nếu có lỗi
        """
        try:
            url = f"{self.config.HUTECH_API_BASE_URL}{self.config.HUTECH_HOC_PHAN_DANH_SACH_SINH_VIEN_ENDPOINT}"
            
            # Tạo headers riêng cho API học phần
            headers = self.config.HUTECH_MOBILE_HEADERS.copy()
            headers["authorization"] = f"JWT {token}"
            
            # Tạo query parameters
            params = {
                "key_lop_hoc_phan": key_lop_hoc_phan
            }
            
            async with aiohttp.ClientSession() as session:
                async with session.get(
                    url,
                    headers=headers,
                    params=params
                ) as response:
                    if response.status == 200:
                        return await response.json()
                    else:
                        error_text = await response.text()
                        logger.error(f"Danh sách sinh viên API error: {response.status} - {error_text}")
                        return {
                            "error": True,
                            "status_code": response.status,
                            "message": error_text
                        }
        
        except aiohttp.ClientError as e:
            logger.error(f"HTTP client error: {e}")
            return {
                "error": True,
                "message": f"Lỗi kết nối: {str(e)}"
            }
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error: {e}")
            return {
                "error": True,
                "message": f"Lỗi phân tích dữ liệu: {str(e)}"
            }
        except Exception as e:
            logger.error(f"Unexpected error: {e}")
            return {
                "error": True,
                "message": f"Lỗi không xác định: {str(e)}"
            }
    
    
    async def _get_user_token(self, telegram_user_id: int) -> Optional[str]:
        """
        Lấy token của người dùng từ database (ưu tiên token từ old_login_info cho các API cũ).
        """
        try:
            response_data = await self.db_manager.get_user_login_response(telegram_user_id)
            if not response_data:
                return None

            # Ưu tiên sử dụng token từ old_login_info cho các API elearning cũ
            old_login_info = response_data.get("old_login_info")
            if isinstance(old_login_info, dict) and old_login_info.get("token"):
                return old_login_info["token"]
            
            # Nếu không, sử dụng token chính
            return response_data.get("token")

        except Exception as e:
            logger.error(f"Error getting token for user {telegram_user_id}: {e}")
            return None
    
    def _process_nam_hoc_hoc_ky_data(self, nam_hoc_hoc_ky_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Xử lý dữ liệu năm học - học kỳ
        
        Args:
            nam_hoc_hoc_ky_data: Dữ liệu năm học - học kỳ thô từ API
            
        Returns:
            Dữ liệu đã được xử lý
        """
        try:
            # Sắp xếp theo mã năm học - học kỳ (mới nhất lên đầu)
            sorted_data = sorted(nam_hoc_hoc_ky_data, key=lambda x: x.get("ma_hoc_ky", ""), reverse=True)
            
            return {
                "nam_hoc_hoc_ky_list": sorted_data
            }
        
        except Exception as e:
            logger.error(f"Error processing năm học - học kỳ data: {e}")
            return {
                "nam_hoc_hoc_ky_list": []
            }
    
    def _process_search_hoc_phan_data(self, search_hoc_phan_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Xử lý dữ liệu tìm kiếm học phần
        
        Args:
            search_hoc_phan_data: Dữ liệu tìm kiếm học phần thô từ API
            
        Returns:
            Dữ liệu đã được xử lý
        """
        try:
            # Sắp xếp theo năm học, học kỳ, tên môn học
            sorted_data = sorted(search_hoc_phan_data, key=lambda x: (
                x.get("json_thong_tin", {}).get("nam_hoc", ""),
                x.get("json_thong_tin", {}).get("hoc_ky", ""),
                x.get("json_thong_tin", {}).get("ten_mon_hoc", "")
            ), reverse=True)
            
            return {
                "hoc_phan_list": sorted_data
            }
        
        except Exception as e:
            logger.error(f"Error processing search học phần data: {e}")
            return {
                "hoc_phan_list": []
            }
    
    def _process_diem_danh_data(self, diem_danh_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Xử lý dữ liệu điểm danh
        
        Args:
            diem_danh_data: Dữ liệu điểm danh thô từ API
            
        Returns:
            Dữ liệu đã được xử lý
        """
        try:
            
            # Hàm chuyển đổi ngày từ chuỗi sang datetime để sắp xếp đúng
            def parse_date(date_str):
                try:
                    # Định dạng ngày là dd/mm/yyyy
                    return datetime.strptime(date_str, "%d/%m/%Y")
                except (ValueError, TypeError):
                    # Nếu không thể chuyển đổi, trả về một ngày rất xa trong tương lai
                    return datetime.max
            
            # Sắp xếp theo ngày học tăng dần (từ cũ đến mới)
            sorted_data = sorted(diem_danh_data, key=lambda x: parse_date(x.get("lich_trinh", {}).get("ngay_hoc", "")))
            
            return {
                "diem_danh_list": sorted_data
            }
        
        except Exception as e:
            logger.error(f"Error processing điểm danh data: {e}", exc_info=True)
            return {
                "diem_danh_list": []
            }
    
    def _process_danh_sach_sinh_vien_data(self, danh_sach_sinh_vien_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Xử lý dữ liệu danh sách sinh viên
        
        Args:
            danh_sach_sinh_vien_data: Dữ liệu danh sách sinh viên thô từ API
            
        Returns:
            Dữ liệu đã được xử lý
        """
        try:
            # Lấy thông tin lớp học phần
            lop_info = danh_sach_sinh_vien_data.get("lop", {})
            json_member = lop_info.get("json_member", {})
            
            # Chuyển đổi json_member thành danh sách sinh viên
            sinh_vien_list = []
            for mssv, info in json_member.items():
                # Tách họ và tên
                ho_ten = info.get("ho_ten", "")
                parts = ho_ten.split()
                if len(parts) > 1:
                    ho = " ".join(parts[:-1])
                    ten = parts[-1]
                else:
                    ho = ""
                    ten = ho_ten
                
                sinh_vien_list.append({
                    "mssv": mssv,
                    "ho": ho,
                    "ten": ten,
                    "lop": info.get("lop", ""),
                    "ho_ten_day_du": ho_ten  # Giữ họ tên đầy đủ để sử dụng nếu cần
                })
            
            # Sắp xếp theo bảng chữ cái tiếng Việt (ưu tiên: Tên trước, Họ sau)
            import locale
            try:
                # Thiết lập locale cho tiếng Việt
                locale.setlocale(locale.LC_COLLATE, 'vi_VN.UTF-8')
                # Sắp xếp theo Tên trước, nếu trùng thì sắp xếp theo Họ
                sinh_vien_list.sort(key=lambda x: (locale.strxfrm(x["ten"]), locale.strxfrm(x["ho"])))
            except locale.Error:
                # Nếu không thể thiết lập locale, sắp xếp theo phương pháp thông thường
                logger.warning("Could not set Vietnamese locale, using default sorting")
                sinh_vien_list.sort(key=lambda x: (x["ten"], x["ho"]))
            
            return {
                "lop_info": lop_info,
                "sinh_vien_list": sinh_vien_list
            }
        
        except Exception as e:
            logger.error(f"Error processing danh sách sinh viên data: {e}")
            return {
                "lop_info": {},
                "sinh_vien_list": []
            }
    
    def format_nam_hoc_hoc_ky_message(self, nam_hoc_hoc_ky_data: Dict[str, Any]) -> str:
        """
        Định dạng dữ liệu năm học - học kỳ thành tin nhắn
        
        Args:
            nam_hoc_hoc_ky_data: Dữ liệu năm học - học kỳ đã được xử lý
            
        Returns:
            Chuỗi tin nhắn đã định dạng
        """
        try:
            nam_hoc_hoc_ky_list = nam_hoc_hoc_ky_data.get("nam_hoc_hoc_ky_list", [])
            timestamp_str = nam_hoc_hoc_ky_data.get("timestamp")

            if not nam_hoc_hoc_ky_list:
                return "📚 *Học Phần*\n\nKhông có dữ liệu năm học - học kỳ."

            message = "📚 *Danh Sách Năm Học - Học Kỳ*\n\n"
            message += "Chọn một hoặc nhiều học kỳ để tìm kiếm học phần.\n\n"
            
            for i, item in enumerate(nam_hoc_hoc_ky_list):
                ma_hoc_ky = item.get("ma_hoc_ky", "N/A")
                ten_hoc_ky = item.get("ten_hoc_ky", "N/A")
                
                message += f"*{i+1}. {ten_hoc_ky}*\n"
                message += f"   - *Mã:* `{ma_hoc_ky}`\n\n"

            if timestamp_str:
                try:
                    ts_utc = datetime.fromisoformat(timestamp_str)
                    ts_local = ts_utc + timedelta(hours=7)
                    message += f"\n_Dữ liệu cập nhật lúc: {ts_local.strftime('%H:%M %d/%m/%Y')}_"
                except (ValueError, TypeError):
                    pass
            
            return message
        
        except Exception as e:
            logger.error(f"Error formatting năm học - học kỳ message: {e}")
            return f"Lỗi định dạng danh sách năm học - học kỳ: {str(e)}"
    
    def format_search_hoc_phan_message(self, search_hoc_phan_data: Dict[str, Any]) -> str:
        """
        Định dạng dữ liệu tìm kiếm học phần thành tin nhắn
        
        Args:
            search_hoc_phan_data: Dữ liệu tìm kiếm học phần đã được xử lý
            
        Returns:
            Chuỗi tin nhắn đã định dạng
        """
        try:
            hoc_phan_list = search_hoc_phan_data.get("hoc_phan_list", [])
            timestamp_str = search_hoc_phan_data.get("timestamp")

            if not hoc_phan_list:
                return "📚 *Kết Quả Tìm Kiếm*\n\nKhông có học phần nào được tìm thấy."
            
            message = "📚 *Kết Quả Tìm Kiếm Học Phần*\n\n"
            
            for i, item in enumerate(hoc_phan_list):
                thong_tin = item.get("json_thong_tin", {})
                ten_mon_hoc = thong_tin.get("ten_mon_hoc", "N/A")
                ma_mon_hoc = thong_tin.get("ma_mon_hoc", "N/A")
                nam_hoc = thong_tin.get("nam_hoc", "N/A")
                hoc_ky = thong_tin.get("hoc_ky", "N/A")
                nhom_hoc = thong_tin.get("nhom_hoc", "N/A")
                so_tc = thong_tin.get("so_tc", "N/A")
                
                message += f"*{i+1}. {ten_mon_hoc}*\n"
                message += f"   - *Mã HP:* `{ma_mon_hoc}`\n"
                message += f"   - *Học kỳ:* `{nam_hoc} - HK{hoc_ky}`\n"
                message += f"   - *Nhóm:* `{nhom_hoc}` | *Số TC:* `{so_tc}`\n\n"
            
            if timestamp_str:
                try:
                    ts_utc = datetime.fromisoformat(timestamp_str)
                    ts_local = ts_utc + timedelta(hours=7)
                    message += f"\n_Dữ liệu cập nhật lúc: {ts_local.strftime('%H:%M %d/%m/%Y')}_"
                except (ValueError, TypeError):
                    pass

            return message
        
        except Exception as e:
            logger.error(f"Error formatting search học phần message: {e}")
            return f"Lỗi định dạng danh sách học phần: {str(e)}"
    
    def format_hoc_phan_detail_message(self, hoc_phan_data: Dict[str, Any]) -> str:
        """
        Định dạng dữ liệu chi tiết học phần thành tin nhắn
        
        Args:
            hoc_phan_data: Dữ liệu học phần đã được xử lý
            
        Returns:
            Chuỗi tin nhắn đã định dạng
        """
        try:
            thong_tin = hoc_phan_data.get("json_thong_tin", {})
            timestamp_str = hoc_phan_data.get("timestamp")
            ten_mon_hoc = thong_tin.get("ten_mon_hoc", "N/A")
            ma_mon_hoc = thong_tin.get("ma_mon_hoc", "N/A")
            nam_hoc = thong_tin.get("nam_hoc", "N/A")
            hoc_ky = thong_tin.get("hoc_ky", "N/A")
            nhom_hoc = thong_tin.get("nhom_hoc", "N/A")
            so_tc = thong_tin.get("so_tc", "N/A")
            nhom_thuc_hanh = thong_tin.get("nhom_thuc_hanh", "")
            
            message = f"📚 *Chi Tiết Học Phần*\n\n"
            message += f"*{ten_mon_hoc}*\n"
            message += f"  - *Mã HP:* `{ma_mon_hoc}`\n"
            message += f"  - *Học kỳ:* `{nam_hoc} - HK{hoc_ky}`\n"
            message += f"  - *Nhóm:* `{nhom_hoc}`\n"
            message += f"  - *Số TC:* `{so_tc}`\n"
            if nhom_thuc_hanh:
                message += f"  - *Nhóm TH:* `{nhom_thuc_hanh}`\n"
            
            if timestamp_str:
                try:
                    ts_utc = datetime.fromisoformat(timestamp_str)
                    ts_local = ts_utc + timedelta(hours=7)
                    message += f"\n_Dữ liệu cập nhật lúc: {ts_local.strftime('%H:%M %d/%m/%Y')}_"
                except (ValueError, TypeError):
                    pass

            return message
        
        except Exception as e:
            logger.error(f"Error formatting học phần detail message: {e}")
            return f"Lỗi định dạng chi tiết học phần: {str(e)}"
    
    def format_diem_danh_message(self, diem_danh_data: Dict[str, Any]) -> str:
        """
        Định dạng dữ liệu điểm danh thành tin nhắn
        
        Args:
            diem_danh_data: Dữ liệu điểm danh đã được xử lý
            
        Returns:
            Chuỗi tin nhắn đã định dạng
        """
        try:
            diem_danh_list = diem_danh_data.get("diem_danh_list", [])
            timestamp_str = diem_danh_data.get("timestamp")

            if not diem_danh_list:
                return "📝 *Lịch Sử Điểm Danh*\n\nKhông có dữ liệu điểm danh."
            
            message = "📝 *Lịch Sử Điểm Danh*\n"
            
            total_sessions = len(diem_danh_list)
            present_sessions = sum(1 for item in diem_danh_list if item and item.get("diem_danh") and item.get("diem_danh", {}).get("ket_qua") == "co_mat")
            absent_sessions = sum(1 for item in diem_danh_list if item and item.get("diem_danh") and item.get("diem_danh", {}).get("ket_qua") == "vang_mat")
            
            message += f"\n*Tổng quan:*\n"
            message += f"  - ✅ *Có mặt:* `{present_sessions}/{total_sessions}`\n"
            message += f"  - ❌ *Vắng mặt:* `{absent_sessions}/{total_sessions}`\n"

            message += "\n- - - - - *Chi Tiết* - - - - -\n"

            for item in diem_danh_list:
                if not item:
                    continue
                lich_trinh = item.get("lich_trinh", {})
                diem_danh = item.get("diem_danh") or {}
                
                ngay_hoc = lich_trinh.get("ngay_hoc", "N/A")
                gio_bat_dau = lich_trinh.get("gio_bat_dau", "N/A")
                gio_ket_thuc = lich_trinh.get("gio_ket_thuc", "N/A")
                ma_phong = lich_trinh.get("ma_phong", "N/A")
                
                ket_qua = diem_danh.get("ket_qua", "chua_diem_danh")
                
                if ket_qua == "co_mat":
                    status_icon = "✅"
                    status_text = "Có mặt"
                elif ket_qua == "vang_mat":
                    status_icon = "❌"
                    status_text = "Vắng mặt"
                else:
                    status_icon = "❔"
                    status_text = "Chưa điểm danh"

                message += f"\n*{ngay_hoc}* ({gio_bat_dau} - {gio_ket_thuc})\n"
                message += f"   - *Trạng thái:* {status_icon} {status_text}\n"
                message += f"   - *Phòng:* `{ma_phong}`\n"
            
            if timestamp_str:
                try:
                    ts_utc = datetime.fromisoformat(timestamp_str)
                    ts_local = ts_utc + timedelta(hours=7)
                    message += f"\n\n_Dữ liệu cập nhật lúc: {ts_local.strftime('%H:%M %d/%m/%Y')}_"
                except (ValueError, TypeError):
                    pass

            return message
        
        except Exception as e:
            logger.error(f"Error formatting điểm danh message: {e}", exc_info=True)
            return f"Lỗi định dạng lịch sử điểm danh: {str(e)}"
    
    def generate_danh_sach_sinh_vien_xlsx(self, danh_sach_sinh_vien_data: Dict[str, Any]) -> io.BytesIO:
        """
        Tạo file Excel danh sách sinh viên
        
        Args:
            danh_sach_sinh_vien_data: Dữ liệu danh sách sinh viên đã được xử lý
            
        Returns:
            File Excel dưới dạng BytesIO
        """
        try:
            lop_info = danh_sach_sinh_vien_data.get("lop_info", {})
            sinh_vien_list = danh_sach_sinh_vien_data.get("sinh_vien_list", [])
            
            # Tạo workbook mới
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Danh sách sinh viên"
            
            # Định dạng tiêu đề
            title_font = Font(name='Arial', size=14, bold=True)
            header_font = Font(name='Arial', size=12, bold=True)
            cell_font = Font(name='Arial', size=11)
            header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            cell_alignment = Alignment(horizontal='left', vertical='center')
            stt_alignment = Alignment(horizontal='center', vertical='center')
            
            # Thêm thông tin lớp học phần
            thong_tin = lop_info.get("json_thong_tin", {})
            ten_mon_hoc = thong_tin.get("ten_mon_hoc", "")
            ma_mon_hoc = thong_tin.get("ma_mon_hoc", "")
            nam_hoc = thong_tin.get("nam_hoc", "")
            hoc_ky = thong_tin.get("hoc_ky", "")
            nhom_hoc = thong_tin.get("nhom_hoc", "")
            
            # Cập nhật merge cells để chứa thêm cột STT
            ws.merge_cells('A1:E1')
            ws['A1'] = f"DANH SÁCH SINH VIÊN LỚP HỌC PHẦN"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('A2:E2')
            ws['A2'] = f"{ten_mon_hoc} ({ma_mon_hoc})"
            ws['A2'].font = header_font
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('A3:E3')
            ws['A3'] = f"Năm học: {nam_hoc} - Học kỳ: {hoc_ky} - Nhóm học: {nhom_hoc}"
            ws['A3'].font = cell_font
            ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Thêm tiêu đề bảng (bao gồm cột STT)
            headers = ['STT', 'MSSV', 'Họ', 'Tên', 'Lớp']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Thêm dữ liệu sinh viên (bao gồm cột STT)
            for row_num, sinh_vien in enumerate(sinh_vien_list, 6):
                # Thêm số thứ tự
                ws.cell(row=row_num, column=1, value=row_num - 5).font = cell_font
                ws.cell(row=row_num, column=1).alignment = stt_alignment
                
                # Thêm MSSV
                ws.cell(row=row_num, column=2, value=sinh_vien["mssv"]).font = cell_font
                ws.cell(row=row_num, column=2).alignment = cell_alignment
                
                # Thêm Họ
                ws.cell(row=row_num, column=3, value=sinh_vien["ho"]).font = cell_font
                ws.cell(row=row_num, column=3).alignment = cell_alignment
                
                # Thêm Tên
                ws.cell(row=row_num, column=4, value=sinh_vien["ten"]).font = cell_font
                ws.cell(row=row_num, column=4).alignment = cell_alignment
                
                # Thêm Lớp
                ws.cell(row=row_num, column=5, value=sinh_vien["lop"]).font = cell_font
                ws.cell(row=row_num, column=5).alignment = cell_alignment
            
            # Điều chỉnh độ rộng cột (bao gồm cột STT)
            ws.column_dimensions['A'].width = 5   # STT
            ws.column_dimensions['B'].width = 15  # MSSV
            ws.column_dimensions['C'].width = 25  # Họ
            ws.column_dimensions['D'].width = 15  # Tên
            ws.column_dimensions['E'].width = 15  # Lớp
            
            # Lưu file vào BytesIO
            file_stream = io.BytesIO()
            wb.save(file_stream)
            file_stream.seek(0)
            
            return file_stream
        
        except Exception as e:
            logger.error(f"Error generating danh sách sinh viên XLSX: {e}")
            raise e
    
    def get_nam_hoc_hoc_ky_list(self, nam_hoc_hoc_ky_data: Dict[str, Any]) -> List[Dict[str, str]]:
        """
        Lấy danh sách năm học - học kỳ để hiển thị trong menu
        
        Args:
            nam_hoc_hoc_ky_data: Dữ liệu năm học - học kỳ đã được xử lý
            
        Returns:
            Danh sách năm học - học kỳ với thông tin hiển thị
        """
        try:
            nam_hoc_hoc_ky_list = nam_hoc_hoc_ky_data.get("nam_hoc_hoc_ky_list", [])
            
            if not nam_hoc_hoc_ky_list:
                return []
            
            result = []
            
            for i, item in enumerate(nam_hoc_hoc_ky_list):
                ma_hoc_ky = item.get("ma_hoc_ky", "")
                ten_hoc_ky = item.get("ten_hoc_ky", "")
                
                result.append({
                    "key": ma_hoc_ky,
                    "name": f"{ten_hoc_ky}",
                    "display": str(i+1)
                })
            
            return result
        
        except Exception as e:
            logger.error(f"Error getting năm học - học kỳ list: {e}")
            return []
    
    def get_hoc_phan_list(self, search_hoc_phan_data: Dict[str, Any]) -> List[Dict[str, str]]:
        """
        Lấy danh sách học phần để hiển thị trong menu
        
        Args:
            search_hoc_phan_data: Dữ liệu tìm kiếm học phần đã được xử lý
            
        Returns:
            Danh sách học phần với thông tin hiển thị
        """
        try:
            hoc_phan_list = search_hoc_phan_data.get("hoc_phan_list", [])
            
            if not hoc_phan_list:
                return []
            
            result = []
            
            for i, item in enumerate(hoc_phan_list):
                thong_tin = item.get("json_thong_tin", {})
                ten_mon_hoc = thong_tin.get("ten_mon_hoc", "")
                ma_mon_hoc = thong_tin.get("ma_mon_hoc", "")
                nam_hoc = thong_tin.get("nam_hoc", "")
                hoc_ky = thong_tin.get("hoc_ky", "")
                nhom_hoc = thong_tin.get("nhom_hoc", "")
                key_check = item.get("key_check", "")
                
                display_name = f"{ten_mon_hoc} ({ma_mon_hoc})"
                if len(display_name) > 40:  # Giới hạn độ dài hiển thị
                    display_name = display_name[:37] + "..."
                
                result.append({
                    "key": key_check,
                    "name": display_name,
                    "full_name": f"{ten_mon_hoc} ({ma_mon_hoc}) - {nam_hoc} - HK{hoc_ky} - NH{nhom_hoc}",
                    "display": str(i+1)
                })
            
            return result
        
        except Exception as e:
            logger.error(f"Error getting học phần list: {e}")
            return []