#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Handler xử lý điểm từ hệ thống HUTECH
"""

import json
import logging
import aiohttp
import io
from typing import Dict, Any, Optional, List
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config.config import Config

logger = logging.getLogger(__name__)

class DiemHandler:
    def __init__(self, db_manager, cache_manager):
        self.db_manager = db_manager
        self.cache_manager = cache_manager
        self.config = Config()
    
    async def handle_diem(self, telegram_user_id: int, hocky_key: Optional[str] = None) -> Dict[str, Any]:
        """
        Xử lý lấy điểm của người dùng
        
        Args:
            telegram_user_id: ID của người dùng trên Telegram
            hocky_key: Mã học kỳ (nếu None, lấy tất cả học kỳ)
            
        Returns:
            Dict chứa kết quả và dữ liệu điểm
        """
        try:
            cache_key = f"diem:{telegram_user_id}"

            # 1. Kiểm tra cache
            cached_result = await self.cache_manager.get(cache_key)
            if cached_result:
                diem_data = cached_result.get("data")
                timestamp = cached_result.get("timestamp")
                
                processed_data = self._process_diem_data(diem_data, hocky_key)
                processed_data["timestamp"] = timestamp

                return {
                    "success": True,
                    "message": "Lấy điểm từ cache thành công",
                    "data": processed_data
                }

            # 2. Nếu cache miss, gọi API
            token = await self._get_user_token(telegram_user_id)
            
            if not token:
                return {
                    "success": False,
                    "message": "Bạn chưa đăng nhập. Vui lòng sử dụng /login để đăng nhập.",
                    "data": None
                }
            
            response_data = await self._call_diem_api(token)
            
            # 3. Lưu vào cache
            if response_data and isinstance(response_data, list):
                await self.cache_manager.set(cache_key, response_data, ttl=86400) # Cache trong 24 giờ
            
            # Kiểm tra kết quả
            if response_data and isinstance(response_data, list):
                # Xử lý dữ liệu điểm
                processed_data = self._process_diem_data(response_data, hocky_key)
                
                # Lấy timestamp từ cache manager để đồng bộ
                cached_data = await self.cache_manager.get(cache_key)
                if cached_data:
                    processed_data["timestamp"] = cached_data.get("timestamp")
                else:
                    processed_data["timestamp"] = datetime.utcnow().isoformat()
                
                return {
                    "success": True,
                    "message": "Lấy điểm thành công (dữ liệu mới)",
                    "data": processed_data
                }
            else:
                return {
                    "success": False,
                    "message": "🚫 *Lỗi*\n\nKhông thể lấy dữ liệu điểm. Vui lòng thử lại sau.",
                    "data": response_data,
                    "show_back_button": True
                }
        
        except Exception as e:
            logger.error(f"Điểm error for user {telegram_user_id}: {e}")
            return {
                "success": False,
                "message": f"🚫 *Lỗi*\n\nĐã xảy ra lỗi khi lấy điểm: {str(e)}",
                "data": None,
                "show_back_button": True
            }
    
    async def _call_diem_api(self, token: str) -> Optional[Dict[str, Any]]:
        """
        Gọi API điểm của HUTECH
        
        Args:
            token: Token xác thực
            
        Returns:
            Response data từ API hoặc None nếu có lỗi
        """
        try:
            url = f"{self.config.HUTECH_API_BASE_URL}{self.config.HUTECH_DIEM_ENDPOINT}"
            
            # Tạo headers riêng cho API điểm
            headers = self.config.HUTECH_MOBILE_HEADERS.copy()
            headers["authorization"] = f"JWT {token}"
            
            async with aiohttp.ClientSession() as session:
                async with session.post(
                    url,
                    headers=headers,
                    json={}  # Request body rỗng theo tài liệu
                ) as response:
                    if response.status == 201:
                        return await response.json()
                    else:
                        error_text = await response.text()
                        logger.error(f"Điểm API error: {response.status} - {error_text}")
                        return {
                            "error": True,
                            "status_code": response.status,
                            "message": error_text
                        }
        
        except aiohttp.ClientError as e:
            logger.error(f"HTTP client error: {e}")
            return {
                "error": True,
                "message": f"Lỗi kết nối: {str(e)}"
            }
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error: {e}")
            return {
                "error": True,
                "message": f"Lỗi phân tích dữ liệu: {str(e)}"
            }
        except Exception as e:
            logger.error(f"Unexpected error: {e}")
            return {
                "error": True,
                "message": f"Lỗi không xác định: {str(e)}"
            }
    
    async def _save_diem_response(self, telegram_user_id: int, response_data: Dict[str, Any]) -> bool:
        """
        Lưu response từ API điểm vào database
        
        Args:
            telegram_user_id: ID của người dùng trên Telegram
            response_data: Dữ liệu response từ API
            
        Returns:
            True nếu lưu thành công, False nếu có lỗi
        """
        try:
            # Phương thức này chưa được implement trong db_manager mới, có thể thêm sau nếu cần
            logger.warning("save_diem_response is not implemented in the new db_manager.")
            return True
        except Exception as e:
            logger.error(f"Error saving điểm response for user {telegram_user_id}: {e}")
            return False
    
    async def _get_user_token(self, telegram_user_id: int) -> Optional[str]:
        """
        Lấy token của người dùng từ database (ưu tiên token từ old_login_info cho các API cũ).
        """
        try:
            response_data = await self.db_manager.get_user_login_response(telegram_user_id)
            if not response_data:
                return None

            # Ưu tiên sử dụng token từ old_login_info cho các API elearning cũ
            old_login_info = response_data.get("old_login_info")
            if isinstance(old_login_info, dict) and old_login_info.get("token"):
                return old_login_info["token"]
            
            # Nếu không, sử dụng token chính
            return response_data.get("token")

        except Exception as e:
            logger.error(f"Error getting token for user {telegram_user_id}: {e}")
            return None
    
    def _process_diem_data(self, diem_data: List[Dict[str, Any]], hocky_key: Optional[str] = None) -> Dict[str, Any]:
        """
        Xử lý dữ liệu điểm
        
        Args:
            diem_data: Dữ liệu điểm thô từ API
            hocky_key: Mã học kỳ cần lọc (nếu None, lấy tất cả)
            
        Returns:
            Dữ liệu đã được xử lý
        """
        try:
            # Nhóm điểm theo học kỳ
            hocky_data = {}
            
            for hocky in diem_data:
                if "nam_hoc_hoc_ky" in hocky:
                    current_hocky_key = hocky["nam_hoc_hoc_ky"]
                    hocky_name = hocky.get("nam_hoc_hoc_ky_name", "")
                    diem_chi_tiet = hocky.get("diem_chi_tiet", [])
                    diem_tich_luy = hocky.get("diem_tich_luy", {})
                    
                    # Sắp xếp điểm chi tiết theo tên học phần
                    diem_chi_tiet.sort(key=lambda x: x.get("ten_hp", ""))
                    
                    hocky_data[current_hocky_key] = {
                        "hocky_name": hocky_name,
                        "diem_chi_tiet": diem_chi_tiet,
                        "diem_tich_luy": diem_tich_luy
                    }
            
            # Nếu có chỉ định học kỳ, chỉ trả về học kỳ đó
            if hocky_key and hocky_key in hocky_data:
                return {
                    "selected_hocky": hocky_key,
                    "hocky_data": {hocky_key: hocky_data[hocky_key]}
                }
            
            return {
                "selected_hocky": None,
                "hocky_data": hocky_data
            }
        
        except Exception as e:
            logger.error(f"Error processing điểm data: {e}")
            return {
                "selected_hocky": None,
                "hocky_data": {}
            }
    
    def format_diem_menu_message(self, diem_data: Dict[str, Any]) -> str:
        """
        Định dạng dữ liệu điểm thành menu chọn học kỳ
        
        Args:
            diem_data: Dữ liệu điểm đã được xử lý
            
        Returns:
            Chuỗi tin nhắn đã định dạng
        """
        try:
            if not diem_data:
                return "📊 *Bảng điểm*\n\nKhông có dữ liệu điểm để hiển thị."

            hocky_data = diem_data.get("hocky_data", {})
            
            if not hocky_data:
                return "📊 *Bảng điểm*\n\nKhông có dữ liệu điểm để hiển thị."
            
            sorted_hocky_keys = sorted(hocky_data.keys(), reverse=True)
            
            message = "📊 *Bảng Điểm Các Học Kỳ*\n\n"
            message += "Chọn một học kỳ để xem chi tiết điểm hoặc xuất file Excel.\n\n"
            
            recent_hocky = sorted_hocky_keys[:3]
            
            for i, hocky_key in enumerate(recent_hocky):
                data = hocky_data[hocky_key]
                hocky_name = data.get("hocky_name", "N/A")
                diem_tich_luy = data.get("diem_tich_luy") or {}
                
                dtb_he4 = diem_tich_luy.get("diem_trung_binh_he_4", "N/A")
                so_tc_dat = diem_tich_luy.get("so_tin_chi_dat", "N/A")
                
                message += f"*{i+1}. {hocky_name}*\n"
                message += f"   - *Điểm TB (Hệ 4):* `{dtb_he4}`\n"
                message += f"   - *Số TC Đạt:* `{so_tc_dat}`\n\n"

            timestamp_str = diem_data.get("timestamp")
            if timestamp_str:
                try:
                    ts_utc = datetime.fromisoformat(timestamp_str)
                    ts_local = ts_utc + timedelta(hours=7)
                    message += f"_Dữ liệu cập nhật lúc: {ts_local.strftime('%H:%M %d/%m/%Y')}_"
                except (ValueError, TypeError):
                    pass
            
            return message
        
        except Exception as e:
            logger.error(f"Error formatting điểm menu message: {e}")
            return f"Lỗi định dạng menu điểm: {str(e)}"
    
    def format_older_hocky_menu_message(self, diem_data: Dict[str, Any]) -> str:
        """
        Định dạng dữ liệu điểm thành menu chọn học kỳ cũ hơn
        
        Args:
            diem_data: Dữ liệu điểm đã được xử lý
            
        Returns:
            Chuỗi tin nhắn đã định dạng
        """
        try:
            if not diem_data:
                return "📊 *Các Học Kỳ Cũ Hơn*\n\nKhông có dữ liệu điểm để hiển thị."

            hocky_data = diem_data.get("hocky_data", {})
            
            if not hocky_data:
                return "📊 *Các Học Kỳ Cũ Hơn*\n\nKhông có dữ liệu điểm để hiển thị."
            
            sorted_hocky_keys = sorted(hocky_data.keys(), reverse=True)
            
            message = "📊 *Các Học Kỳ Cũ Hơn*\n\n"
            message += "Chọn một học kỳ để xem chi tiết điểm hoặc xuất file Excel.\n\n"
            
            older_hocky = sorted_hocky_keys[3:]
            
            for i, hocky_key in enumerate(older_hocky):
                data = hocky_data[hocky_key]
                hocky_name = data.get("hocky_name", "N/A")
                diem_tich_luy = data.get("diem_tich_luy") or {}
                
                dtb_he4 = diem_tich_luy.get("diem_trung_binh_he_4", "N/A")
                so_tc_dat = diem_tich_luy.get("so_tin_chi_dat", "N/A")
                
                message += f"*{i+1}. {hocky_name}*\n"
                message += f"   - *Điểm TB (Hệ 4):* `{dtb_he4}`\n"
                message += f"   - *Số TC Đạt:* `{so_tc_dat}`\n\n"

            timestamp_str = diem_data.get("timestamp")
            if timestamp_str:
                try:
                    ts_utc = datetime.fromisoformat(timestamp_str)
                    ts_local = ts_utc + timedelta(hours=7)
                    message += f"_Dữ liệu cập nhật lúc: {ts_local.strftime('%H:%M %d/%m/%Y')}_"
                except (ValueError, TypeError):
                    pass
            
            return message
        
        except Exception as e:
            logger.error(f"Error formatting older học kỳ menu message: {e}")
            return f"Lỗi định dạng menu điểm học kỳ cũ: {str(e)}"

    def format_diem_detail_message(self, diem_data: Dict[str, Any]) -> str:
        """
        Định dạng dữ liệu điểm chi tiết của một học kỳ
        
        Args:
            diem_data: Dữ liệu điểm đã được xử lý
            
        Returns:
            Chuỗi tin nhắn đã định dạng
        """
        try:
            hocky_data = diem_data.get("hocky_data", {})
            selected_hocky = diem_data.get("selected_hocky")
            
            if not hocky_data or not selected_hocky or selected_hocky not in hocky_data:
                return "📊 Không có dữ liệu điểm chi tiết."
            
            data = hocky_data[selected_hocky]
            hocky_name = data.get("hocky_name", "N/A")
            diem_chi_tiet = data.get("diem_chi_tiet", [])
            diem_tich_luy = data.get("diem_tich_luy", {})
            
            message = f"📊 *Điểm Chi Tiết - {hocky_name}*\n"
            
            if diem_tich_luy:
                dtb_he4 = diem_tich_luy.get("diem_trung_binh_he_4", "N/A")
                dtb_tl_he4 = diem_tich_luy.get("diem_trung_binh_tich_luy_he_4", "N/A")
                so_tc_dat = diem_tich_luy.get("so_tin_chi_dat", "N/A")
                so_tc_tl = diem_tich_luy.get("so_tin_chi_tich_luy", "N/A")
                
                message += "\n*Tổng Kết Học Kỳ:*\n"
                message += f"  - *Điểm TB (Hệ 4):* `{dtb_he4}`\n"
                message += f"  - *Điểm TB Tích Lũy (Hệ 4):* `{dtb_tl_he4}`\n"
                message += f"  - *Số TC Đạt:* `{so_tc_dat}`\n"
                message += f"  - *Tổng TC Tích Lũy:* `{so_tc_tl}`\n"
            
            if diem_chi_tiet:
                message += "\n- - - - - *Điểm Môn Học* - - - - -\n"
                
                for mon in diem_chi_tiet:
                    ten_hp = mon.get("ten_hp", "N/A")
                    ma_hp = mon.get("ma_hp", "N/A")
                    stc = mon.get("stc", "N/A")
                    diem_he10 = mon.get("diem_he_10", "N/A")
                    diem_he4 = mon.get("diem_he_4", "N/A")
                    diem_chu = mon.get("diem_chu", "N/A")
                    
                    message += f"\n📚 *{ten_hp}*\n"
                    message += f"   - *Mã HP:* `{ma_hp}`\n"
                    message += f"   - *Số TC:* `{stc}`\n"
                    message += f"   - *Điểm Tổng Kết:* `{diem_he10}` (Hệ 10) - `{diem_he4}` (Hệ 4) - `{diem_chu}` (Điểm chữ)\n"
                    
                    diem_kt1 = mon.get("diem_kiem_tra_1", "")
                    diem_kt2 = mon.get("diem_kiem_tra_2", "")
                    diem_thi = mon.get("diem_thi", "")
                    
                    if diem_kt1 or diem_kt2 or diem_thi:
                        components = []
                        if diem_kt1: components.append(f"KT1: `{diem_kt1}`")
                        if diem_kt2: components.append(f"KT2: `{diem_kt2}`")
                        if diem_thi: components.append(f"Thi: `{diem_thi}`")
                        message += f"   - *Điểm thành phần:* {', '.join(components)}\n"
            else:
                message += "\nKhông có điểm chi tiết trong học kỳ này.\n"

            timestamp_str = diem_data.get("timestamp")
            if timestamp_str:
                try:
                    ts_utc = datetime.fromisoformat(timestamp_str)
                    ts_local = ts_utc + timedelta(hours=7)
                    message += f"\n\n_Dữ liệu cập nhật lúc: {ts_local.strftime('%H:%M %d/%m/%Y')}_"
                except (ValueError, TypeError):
                    pass
            
            return message
        
        except Exception as e:
            logger.error(f"Error formatting điểm detail message: {e}")
            return f"Lỗi định dạng điểm chi tiết: {str(e)}"
    
    def get_hocky_list(self, diem_data: Dict[str, Any]) -> List[Dict[str, str]]:
        """
        Lấy danh sách học kỳ để hiển thị trong menu
        
        Args:
            diem_data: Dữ liệu điểm đã được xử lý
            
        Returns:
            Danh sách học kỳ với thông tin hiển thị
        """
        try:
            hocky_data = diem_data.get("hocky_data", {})
            
            if not hocky_data:
                return []
            
            # Sắp xếp học kỳ theo mã (mới nhất lên đầu)
            sorted_hocky_keys = sorted(hocky_data.keys(), reverse=True)
            
            result = []
            
            # Thêm 3 học kỳ gần nhất
            for i, hocky_key in enumerate(sorted_hocky_keys[:3]):
                data = hocky_data[hocky_key]
                hocky_name = data.get("hocky_name", "")
                
                result.append({
                    "key": hocky_key,
                    "name": f"{hocky_name}",
                    "display": str(i+1)
                })
            
            # Nếu có nhiều hơn 3 học kỳ, thêm nút "Xem thêm"
            if len(sorted_hocky_keys) > 3:
                result.append({
                    "key": "more",
                    "name": "Xem thêm học kỳ cũ hơn",
                    "display": "4"
                })
            
            return result
        
        except Exception as e:
            logger.error(f"Error getting học kỳ list: {e}")
            return []
    
    def get_older_hocky_list(self, diem_data: Dict[str, Any]) -> List[Dict[str, str]]:
        """
        Lấy danh sách học kỳ cũ hơn để hiển thị
        
        Args:
            diem_data: Dữ liệu điểm đã được xử lý
            
        Returns:
            Danh sách học kỳ cũ với thông tin hiển thị
        """
        try:
            hocky_data = diem_data.get("hocky_data", {})
            
            if not hocky_data:
                return []
            
            # Sắp xếp học kỳ theo mã (mới nhất lên đầu)
            sorted_hocky_keys = sorted(hocky_data.keys(), reverse=True)
            
            # Lấy các học kỳ cũ hơn (từ vị trí thứ 3 trở đi)
            older_hocky_keys = sorted_hocky_keys[3:]
            
            result = []
            
            for i, hocky_key in enumerate(older_hocky_keys):
                data = hocky_data[hocky_key]
                hocky_name = data.get("hocky_name", "")
                
                result.append({
                    "key": hocky_key,
                    "name": f"{hocky_name}",
                    "display": str(i+1)
                })
            
            return result
        
        except Exception as e:
            logger.error(f"Error getting older học kỳ list: {e}")
            return []

    def generate_diem_xlsx(self, diem_data: Dict[str, Any], hocky_key: Optional[str] = None) -> io.BytesIO:
        """
        Tạo file Excel điểm
        
        Args:
            diem_data: Dữ liệu điểm đã được xử lý
            hocky_key: Mã học kỳ (nếu None, xuất toàn bộ)
            
        Returns:
            File Excel dưới dạng BytesIO
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Định dạng
            title_font = Font(name='Arial', size=16, bold=True)
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            cell_font = Font(name='Arial', size=11)
            tich_luy_font = Font(name='Arial', size=11, bold=True)
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Lấy dữ liệu
            hocky_data = diem_data.get("hocky_data", {})
            
            if hocky_key and hocky_key in hocky_data:
                # Xuất điểm của một học kỳ
                data = hocky_data[hocky_key]
                hocky_name = data.get("hocky_name", "")
                ws.title = hocky_name
                
                self._write_hocky_to_sheet(ws, hocky_name, data, title_font, header_font, cell_font, tich_luy_font, header_fill, center_alignment, left_alignment, thin_border)
            else:
                # Xuất điểm toàn bộ
                ws.title = "Điểm Toàn Bộ"
                
                # Sắp xếp học kỳ theo mã (cũ nhất lên đầu để xuất file)
                sorted_hocky_keys = sorted(hocky_data.keys())
                
                current_row = 1
                for key in sorted_hocky_keys:
                    data = hocky_data[key]
                    hocky_name = data.get("hocky_name", "")
                    
                    current_row = self._write_hocky_to_sheet(ws, hocky_name, data, title_font, header_font, cell_font, tich_luy_font, header_fill, center_alignment, left_alignment, thin_border, start_row=current_row)
                    current_row += 2 # Thêm khoảng cách giữa các học kỳ

            # Lưu file vào BytesIO
            file_stream = io.BytesIO()
            wb.save(file_stream)
            file_stream.seek(0)
            
            return file_stream
        
        except Exception as e:
            logger.error(f"Error generating điểm XLSX: {e}", exc_info=True)
            raise e

    def _write_hocky_to_sheet(self, ws, hocky_name, data, title_font, header_font, cell_font, tich_luy_font, header_fill, center_alignment, left_alignment, thin_border, start_row=1):
        """
        Ghi dữ liệu điểm của một học kỳ vào sheet
        """
        # Tiêu đề học kỳ
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
        cell = ws.cell(row=start_row, column=1, value=f"BẢNG ĐIỂM HỌC KỲ: {hocky_name.upper()}")
        cell.font = title_font
        cell.alignment = center_alignment
        
        # Tiêu đề bảng
        headers = ["STT", "Mã HP", "Tên học phần", "STC", "KT1", "KT2", "Thi", "Điểm 10", "Điểm 4", "Điểm chữ"]
        header_row = start_row + 1
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Dữ liệu điểm chi tiết
        diem_chi_tiet = data.get("diem_chi_tiet", [])
        current_row = header_row + 1
        for i, mon in enumerate(diem_chi_tiet, 1):
            ws.cell(row=current_row, column=1, value=i).alignment = center_alignment
            ws.cell(row=current_row, column=2, value=mon.get("ma_hp", "")).alignment = left_alignment
            ws.cell(row=current_row, column=3, value=mon.get("ten_hp", "")).alignment = left_alignment
            ws.cell(row=current_row, column=4, value=mon.get("stc", "")).alignment = center_alignment
            ws.cell(row=current_row, column=5, value=mon.get("diem_kiem_tra_1", "")).alignment = center_alignment
            ws.cell(row=current_row, column=6, value=mon.get("diem_kiem_tra_2", "")).alignment = center_alignment
            ws.cell(row=current_row, column=7, value=mon.get("diem_thi", "")).alignment = center_alignment
            ws.cell(row=current_row, column=8, value=mon.get("diem_he_10", "")).alignment = center_alignment
            ws.cell(row=current_row, column=9, value=mon.get("diem_he_4", "")).alignment = center_alignment
            ws.cell(row=current_row, column=10, value=mon.get("diem_chu", "")).alignment = center_alignment
            
            # Áp dụng font và border
            for col in range(1, 11):
                ws.cell(row=current_row, column=col).font = cell_font
                ws.cell(row=current_row, column=col).border = thin_border

            current_row += 1

        # Dữ liệu điểm tích lũy
        diem_tich_luy = data.get("diem_tich_luy", {})
        if diem_tich_luy:
            tich_luy_data = [
                ("Điểm TB học kỳ (hệ 4)", diem_tich_luy.get("diem_trung_binh_he_4", "")),
                ("Điểm TB tích lũy (hệ 4)", diem_tich_luy.get("diem_trung_binh_tich_luy_he_4", "")),
                ("Số TC đạt", diem_tich_luy.get("so_tin_chi_dat", "")),
                ("Tổng TC tích lũy", diem_tich_luy.get("so_tin_chi_tich_luy", "")),
            ]
            
            for i, (label, value) in enumerate(tich_luy_data):
                ws.merge_cells(start_row=current_row + i, start_column=1, end_row=current_row + i, end_column=3)
                cell_label = ws.cell(row=current_row + i, column=1, value=label)
                cell_label.font = tich_luy_font
                cell_label.alignment = left_alignment
                
                cell_value = ws.cell(row=current_row + i, column=4, value=value)
                cell_value.font = tich_luy_font
                cell_value.alignment = center_alignment

        # Điều chỉnh độ rộng cột
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 5
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 10
        
        return current_row + len(tich_luy_data) if diem_tich_luy else current_row